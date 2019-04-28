import csv
from openpyxl import Workbook
from openpyxl import load_workbook

def load_worksheet(file, sheet_name):
    '''
    Purpose: load an excel file we want to update
        file:(str) xlsx file already created file to modify
        sheet_name:(str) sheet name of file
    '''
    wb = load_workbook(filename = file)
    ws = wb.get_sheet_by_name(sheet_name)
    return ws

def update_row(ws,col_ind,row_ind,val_fill):
    '''
    Purpose: Update an excel file with string, cell by cell. 
    Inputs:
        ws(worksheet): worksheet we want to modify
        col_ind:(int) excel file's column index
        row_ind:(int) excel file's row index
        val_fill(str) value to be filled in the cell
    '''

    ws.cell(row=row_ind,column=col_ind).value = val_fill
    return ws
    
def save_file(wb,file):
    '''
    Purpose: Save the modified excel 
    Inputs:
        wb(workbook): workbook we want to save
        file:(str) xlsx file 
    '''
    wb.save(file)
    return None
    